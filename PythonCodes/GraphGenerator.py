from openpyxl import (load_workbook, Workbook)
from openpyxl.chart import( ScatterChart, Reference, Series)
import math

wb = load_workbook('AdjustedCData.xlsx')

## Names for different worksheets

seseds = wb[wb.sheetnames[0]]
msncodes = wb[wb.sheetnames[1]]

graphs = wb.create_sheet("Graphs")

numberof4s = 0

for k in range(0, 2332):

    print(k)

    if (k%4==0):

        numberof4s+=1

    chart = ScatterChart()
    chart.style = 13

    ## Defines msn code and state for kth 50 chunk

    c = seseds.cell(row=(50*k+2), column=1)
    d = seseds.cell(row=(50*k+2), column=2)

    code = c.value
    state = d.value

    ## Create data coordinates for chart and appends data to chart

    year = Reference(seseds, min_col=3, min_row = (50*k+2), max_row=(50*(k+1)+1))
    data = Reference(seseds, min_col=4, min_row = (50*k+2), max_row=(50*(k+1)+1))
    s = Series(data, xvalues=year)
    chart.append(s)

    ## Collects corresponding despriction and units from msncodes

    title = msncodes.cell(row=(numberof4s+1), column=2)
    y_axis = msncodes.cell(row=(numberof4s+1), column=3)

    ## Defines new title that will Description followed by state

    newtitle = (title.value + " " + state)

    ## Gives appropriate names to title, x-axis and y-axis for each chart

    chart.title = newtitle
    chart.x_axis.title = 'Years'
    chart.y_axis.title = y_axis.value

    ## Spaces charts out appropriately to fit two per page

    numberforA = 17*k+1

    chartposition = "A" + str(numberforA)

    graphs.add_chart(chart, chartposition)

wb.save("Test.xlsx")
