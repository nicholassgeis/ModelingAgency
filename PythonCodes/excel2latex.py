from openpyxl import load_workbook

##def col(n):
##    c=[]
##    while(n>0):
##        c.insert(0,int(n%26))
##        n=(n-c[0])/26
##    s=""
##    for k in c:
##        s+=chr(k+64)
##    return(s)
##
##def cell(ws,x,y):
##    s=str(ws[col(x+1)+str(y+1)])
##    return(s)
##
##def tex(file,x,y):
##    wb = load_workbook(file)
##    ws = wb.active
##    s="\\begin{tabular}\n"
##    for j in range(y):
##        for i in range(x-1):
##            s+=cell(ws,i,j)+" & "
##        s+=cell(ws,x-1,j)+" \\\\ \n"
##    s+="\\end{tabular}"
##    return(s)

def tex(file,x,y):
    wb = load_workbook(file)
    ws = wb.active
    s="\\begin{tabular}\n"
    for j in range(y):
        for i in range(x-1):
            s+=ws.cell(i,j)+" & "
        s+=ws.cell(x-1,j)+" \\\\ \n"
    s+="\\end{tabular}"
    return(s)

print(tex("ProblemCData.xlsx",10,4))
