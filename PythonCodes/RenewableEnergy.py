from openpyxl import (load_workbook, Workbook)
import math

wb = Workbook()

dictionary = ['solar', 'geothermal', 'wind', 'biomass', 'hydroelectric', 'sun', 'Biomass', 'Geothermal', 'Solar', 'Wind', 'Hydroelectric', 'Sunlight']

ws=wb.active
ws.title = "Renewable msncodes"

wbOG = load_workbook('ProblemCData.xlsx')
msncodes = wbOG[wbOG.sheetnames[1]]

i = 1

for item in dictionary:

    for rows in msncodes:    

        for word in rows[1].value.split():
            
            if (word == item):


                ws['A' + str(i)] = rows[0].value
                ws['B' + str(i)] = rows[1].value

                i += 1

wb.save("RenewableEnergyData.xlsx")
