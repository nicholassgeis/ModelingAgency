#!/usr/bin/env python2.7
from openpyxl import (load_workbook, Workbook)
import math

wb = load_workbook('ProblemCData.xlsx', read_only=True)

unitslist = []
msncodes = wb[wb.sheetnames[1]]

for row in msncodes:
    if (row[2].value).lstrip('u') in unitslist:
        continue
    else:
        unitslist.append((row[2].value).lstrip('u'))

print(unitslist)
