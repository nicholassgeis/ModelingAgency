from openpyxl import(load_workbook, Workbook)
import math

## Creates new workbook that will be written to

wb = Workbook()

## Loads Data file that we will change into new workbook

wbOG = load_workbook('ProblemCData.xlsx')

## Names desired worksheet from Data file

seseds = wbOG[wbOG.sheetnames[0]]

## Creates and names worksheet in new file

ws = wb.active
ws.title = "seseds"

ws2 = wb.create_sheet("outliers", 1)
ws2.title = "Date Outliers"

prev_row = seseds[2]

counter = 2

j = 1

for rows in seseds:

    if ( rows[2].value == 'Year' ):

        ws['A1'] = rows[0].value
        ws['B1'] = rows[1].value
        ws['C1'] = rows[2].value
        ws['D1'] = rows[3].value

    elif ((rows[1].value == prev_row[1].value) or (str(rows[2].value) == '1960')):

        ws['A' + str(counter)] = rows[0].value
        ws['B' + str(counter)] = rows[1].value
        ws['C' + str(counter)] = rows[2].value
        ws['D' + str(counter)] = rows[3].value

        counter += 1

    else:

        ws2['A' + str(j)] = rows[0].value
        ws2['B' + str(j)] = rows[2].value

        j += 1

        for i in range(0, int(rows[2].value) - 1960):

            ws['A' + str(counter)] = rows[0].value
            ws['B' + str(counter)] = rows[1].value
            ws['C' + str(counter)] = 1960+i
            ws['D' + str(counter)] = 0

            counter += 1

        ws['A' + str(counter)] = rows[0].value
        ws['B' + str(counter)] = rows[1].value
        ws['C' + str(counter)] = rows[2].value
        ws['D' + str(counter)] = rows[3].value

        counter += 1

    prev_row = rows

wb.save("AdjustedCData.xlsx")

            
