#!/usr/bin/env python2.7
from openpyxl import (load_workbook, Workbook)
import math

wb = load_workbook('AdjustedCData.xlsx', read_only=True)

seseds = wb[wb.sheetnames[0]]
msncodes = wb[wb.sheetnames[1]]
missingdata = []

for row_msn in msncodes:
    print(row_msn[0].value)
    msncode = row_msn[0].value
    counter = 0
    for row_ses in seseds:
        if row_ses[0].value==row_msn[0].value:
            counter += 1
            if counter == 200:
                break
    if counter != 200:
        missingdata.append(row_msn[0].value)

print(missingdata)
